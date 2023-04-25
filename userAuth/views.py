from copy import copy
from datetime import datetime
import json
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth import login, authenticate
from django.views.decorators.csrf import csrf_exempt
from django.utils.html import format_html
from .forms import NewTaskForm, SignUpForm
from .models import Comment, Intern, Mentor, Task, User
import pandas as pd
import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.cell_range import CellRange

def home(request):
    content = {}
    if request.user.is_authenticated:
        interns_list=Intern.objects.filter(mentorid_id=request.user.id)
        
        if request.user.role == 'Mentor':
            pending = Task.objects.filter(mentor_id=request.user.id,progress_status__iexact='To-Do').order_by('-date_updated','-id')
            in_progress = Task.objects.filter(mentor_id=request.user.id,progress_status__iexact='In-Progress').order_by('-date_updated','-id')
            completed = Task.objects.filter(mentor_id=request.user.id,progress_status__iexact='Completed').order_by('-date_updated','-id')
        else:
            pending = Task.objects.filter(internid_id=request.user.id,progress_status__iexact='To-Do').order_by('-date_updated','-id')
            in_progress = Task.objects.filter(internid_id=request.user.id,progress_status__iexact='In-Progress').order_by('-date_updated','-id')
            completed = Task.objects.filter(internid_id=request.user.id,progress_status__iexact='Completed').order_by('-date_updated','-id')
        
        # retrieve task count 
        count_p = pending.count()
        count_i = in_progress.count()
        count_c = completed.count()
        
        content = {'showb':True,'to_do':pending,'in_progress':in_progress,'completed':completed,'interns':interns_list,'pcount':count_p,'ccount':count_c,'icount':count_i}
    return render(request, 'auth/widgets/main.html',content)

@csrf_exempt
def add_tasks(request):
    if request.method == 'POST':
        selectedIDs = json.loads(request.POST.get('data'))
        interns_list=Intern.objects.filter(mentorid_id=request.user.id).values_list('pk', flat=True)

        intern_list = list(interns_list)
        if set(selectedIDs).issubset(set(intern_list)) is False:
            response = JsonResponse({"error": "Invalid Interns. Try Again!"})
            response.status_code = 403
            return response

        file =request.FILES.get('sheet', None)
        print("addtasks",selectedIDs,file)

        if file.name.endswith('.xlsx'):
            df = pd.read_excel(file)
            # forward fill for NaN values
            df = df.fillna(method='ffill',axis=0)

            try:
                df['Intern ID'] = df['Intern ID'].astype(int)
                interns = df['Intern ID'].unique().tolist()

                res_df = df[df['Intern ID'].isin(selectedIDs)]

                wb = openpyxl.load_workbook(file)
                ws = wb.active
                
                # unmerge cells and copy cell values
                all_merged_cell_ranges: list[CellRange] = list(ws.merged_cells.ranges)

                for merged_cell_range in all_merged_cell_ranges:
                    merged_cell: Cell = merged_cell_range.start_cell
                    ws.unmerge_cells(range_string=merged_cell_range.coord)

                    for row_index, col_index in merged_cell_range.cells:
                        cell: Cell = ws.cell(row=row_index, column=col_index)
                        cell.value = merged_cell.value

                        # copy cell styling
                        cell.alignment = copy(merged_cell.alignment)
                        cell.border = copy(merged_cell.border)
                        cell.font = copy(merged_cell.font)


                description = ""
                last_updated_by_id = request.user.id

                # extract source links
                for i in range(2,df.shape[0]+2):
                    if ws.cell(row=i, column=1).value in selectedIDs:
                        links = None
                        due_date = None
                        try:
                            description = ws.cell(row=i, column=4).value
                            links = ws.cell(row=i, column=4).hyperlink.target
                            due_date = ws.cell(row=i, column=3).value.date()
                            progress_status = ws.cell(row=i, column=2).value                      
                            internid_id = ws.cell(row=i, column=1).value
                        except:  
                            description = ws.cell(row=i, column=4).value                
                            dd = ws.cell(row=i, column=3).value
                            if dd is not None:
                                due_date = dd.date()
                            progress_status = ws.cell(row=i, column=2).value
                            internid_id = ws.cell(row=i, column=1).value
                        finally:
                            started_date = completed_date = None
                            completion_status = False
                            if progress_status != 'To Do':
                                started_date = datetime.now().date()
                            if progress_status == 'Done':
                                progress_status ='completed'
                                completed_date = datetime.now().date()
                                completion_status = True
                            else:
                                if progress_status == 'To Do':
                                    progress_status = 'To-do'
                                else:
                                    progress_status = 'In-Progress'

                            print(description,due_date,progress_status,internid_id,started_date,completed_date,completion_status,last_updated_by_id)
                            if links is not None:
                                description=format_html('<a href="{}">{}</a>', links, description)   
                                                     
                            task = Task(description=description, started_date=started_date, 
                                        completed_date=completed_date,due_date=due_date,
                                        completion_status=completion_status, progress_status=progress_status,
                                        internid_id=internid_id, last_updated_by_id=last_updated_by_id, 
                                        mentor_id=last_updated_by_id)
                            task.save()
                                                     
                return JsonResponse({'interns':interns})

            except Exception as e:
                print(f'{e}')

        # if file.name.endswith('.csv'):
        #     try:
        #         df = pd.read_csv(file)
        #         # forward fill for NaN values
        #         df = df.fillna(method='ffill',axis=0)

        #         df['Intern ID'] = df['Intern ID'].astype(int)
        #         interns = df['Intern ID'].unique().tolist()

        #         print(interns, selectedIDs)

        #         res_df = df[df['Intern ID'].isin(selectedIDs)]
        #         res_df.to_excel('out.xlsx',index=None,header=True)

        #         wb = openpyxl.load_workbook('out.xlsx')
        #         ws = wb.active

        #         description = progress_status = ""
        #         internid_id = None
        #         last_updated_by_id = request.user.id

        #         # extract source links
        #         for i in range(2,res_df.shape[0]+2):
        #             links = None
        #             due_date = None
        #             try:
        #                 description = ws.cell(row=i, column=4).value
        #                 links = ws.cell(row=i, column=4).hyperlink.target
        #                 due_date = ws.cell(row=i, column=3).value
        #                 progress_status = ws.cell(row=i, column=2).value                      
        #                 internid_id = ws.cell(row=i, column=1).value
        #             except:  
        #                 description = ws.cell(row=i, column=4).value                
        #                 due_date = ws.cell(row=i, column=3).value
        #                 progress_status = ws.cell(row=i, column=2).value
        #                 internid_id = ws.cell(row=i, column=1).value
        #             finally:
        #                 started_date = completed_date = None
        #                 completion_status = False
        #                 if progress_status != 'To Do':
        #                     started_date = datetime.now().date()
        #                 if progress_status == 'Done':
        #                     progress_status ='completed'
        #                     completed_date = datetime.now().date()
        #                     completion_status = True
        #                 else:
        #                     if progress_status == 'To Do':
        #                         progress_status = 'To-do'
        #                     else:
        #                         progress_status = 'In-Progress'

        #                 print(description,due_date,progress_status,internid_id,started_date,completed_date,completion_status,last_updated_by_id)
        #                 if links is not None:
        #                     description=format_html('<a href="{}">{}</a>', links, description)   
                                                     
        #                 # task = Task(description=description, started_date=started_date, 
        #                 #             completed_date=completed_date,due_date=due_date,
        #                 #             completion_status=completion_status, progress_status=progress_status,
        #                 #             internid_id=internid_id, last_updated_by_id=last_updated_by_id, 
        #                 #             mentor_id=last_updated_by_id)
        #                 # task.save()
                                                     
        #         return JsonResponse({'interns':interns})

        #     except Exception as e:
        #         print(f'{e}')

        return JsonResponse('access',safe=False)

@csrf_exempt
def upload_file(request):
    if request.method == "POST":
        file = request.FILES['sheet']

        # xlsx files
        if file.name.endswith('.xlsx'):
            df = pd.read_excel(file)
            # forward fill for NaN values
            df = df.fillna(method='ffill',axis=0)

            try:
                df['Intern ID'] = df['Intern ID'].astype(int)

                interns = df['Intern ID'].unique().tolist()

                print(type(df),df.columns)
                print(df.shape[0],interns)
                return JsonResponse({'interns':interns})

            except Exception as e:
                print(f'{e}')
        
        # upload csv files
        # if file.name.endswith('.csv'):
        #     try:
        #         df = pd.read_csv(file)
        #         # forward fill for NaN values
        #         df = df.fillna(method='ffill',axis=0)

        #         df['Intern ID'] = df['Intern ID'].astype(int)

        #         interns = df['Intern ID'].unique().tolist()

        #         print(df,df.columns)
        #         print(df.shape[0],interns)
        #         return JsonResponse({'interns':interns})

        #     except Exception as e:
        #         print(f'{e}')            

        return HttpResponse('File upload success',status=204)

@csrf_exempt
def update_tasks(request):
    if request.method == "POST":
        changes = json.loads(request.POST.get("changes"))

        for key in changes.keys():
            task=get_object_or_404(Task,id=key)
            if changes[key]['end_category'] == 'in-progress-tasks':
                task.started_date=datetime.now().date()
                task.progress_status='In-Progress'

            elif changes[key]['end_category'] == 'completed-tasks':
                if task.started_date is None:
                    task.started_date=datetime.now().date()

                task.completed_date=datetime.now().date()
                task.completion_status=True
                task.progress_status='completed'

            if changes[key]['end_category'] != 'pending-tasks':
                task.last_updated_by_id=request.user.id
                task.save()

        response_data = {'status': 'success', 'message': 'Tasks updated successfully'}
        return JsonResponse(response_data)
    else:
        return JsonResponse({"error": "Invalid request method."})

def my_profile(request):
    if request.user.role == 'Mentor':
        user=Mentor.objects.get(mentorid_id=request.user.id)
    else:
        user=Intern.objects.get(internid_id=request.user.id)
    return render(request,'auth/widgets/profile.html',{'collab':user})

@csrf_exempt
def new_comment(request):
    if request.method == 'POST':
        content = request.POST.get('comment')
        task = request.POST.get('task')
        commenter = request.user.id

        comment = Comment(comment=content,commenter_id=commenter,task_id=task)
        comment.save()
    return redirect('home')

def task_details(request,pk):
    # retrieve comment history
    task=get_object_or_404(Task,id=pk)
    object=Comment.objects.filter(task_id=pk).exists()
    comment=Comment.objects.filter(task_id=pk) if object else None
    context={'task':task,'comment':comment}
    return render(request, 'auth/widgets/task_details.html',context)
    
def reset(request):
    return render(request, 'auth/widgets/reset.html')

@csrf_exempt
def update(request):
    if request.user.role == 'intern':
        intern=get_object_or_404(Intern,pk=request.user.id)
        intern.firstname=request.POST.get('firstname')
        intern.lastname=request.POST.get('lastname')        
        intern.college=request.POST.get('college')
        intern.phone=request.POST.get('phone')
        intern.save()

    if request.user.role == 'Mentor':
        mentor=get_object_or_404(Mentor,pk=request.user.id)
        mentor.firstname=request.POST.get('firstname')
        mentor.lastname=request.POST.get('lastname')
        mentor.phone=request.POST.get('phone')
        mentor.save()
    return redirect('profile')

@csrf_exempt
def new_password(request):
    if request.user.is_authenticated:
        user = User.objects.get(id=request.user.id)
        user.set_password(request.POST.get('password-verify'))
        user.save()
        print(request.POST.get('password-verify'),user.first_name)
    return redirect('home')

def name_api(request):
    id = request.GET.get('uuid')
    intern = Intern.objects.get(internid_id=id)
    data = {
        'name': f'{intern.firstname} {intern.lastname}'
    }
    return JsonResponse(data)

def intern_filter(request):
    to_do = Task.objects.filter(mentor_id=request.user.id,progress_status__iexact='To-Do').order_by('-date_updated','-id')
    in_progress = Task.objects.filter(mentor_id=request.user.id,progress_status__iexact='In-Progress').order_by('-date_updated','-id')
    completed = Task.objects.filter(mentor_id=request.user.id,progress_status__iexact='Completed').order_by('-date_updated','-id')

    interns_list=Intern.objects.filter(mentorid_id=request.user.id)
    category = request.GET.get('id')
    if category:
        to_do = to_do.filter(internid_id=category)
        in_progress = in_progress.filter(internid_id=category)
        completed = completed.filter(internid_id=category)
    count_p = to_do.count()
    count_i = in_progress.count()
    count_c = completed.count()
    content={'showb':False,'to_do': to_do,'in_progress': in_progress, 'completed': completed, 'interns':interns_list, 'pcount':count_p,'ccount':count_c,'icount':count_i}
    return render(request, 'auth/widgets/main.html', content)

def intern_details(request):
    usee=request.user.id
    interns_list=Intern.objects.filter(mentorid_id=usee)
    return render(request, 'auth/widgets/tables.html',{'intern_obj':interns_list})

def new_task(request):
    if request.method == 'POST':
        form = NewTaskForm(request.POST,user=request.user.id,role=request.user.role)
        if form.is_valid():
            task=form.save(commit=False)
            if request.user.role == 'Mentor':
                task.internid_id=form.cleaned_data.get('intern')
                task.mentor_id=request.user.id
            else:
                task.internid_id = request.user.id
                task.mentor_id = Intern.objects.get(internid = request.user.id).mentorid_id
            if form.cleaned_data.get('progress_status') == 'completed':
                task.completion_status = True
                task.started_date = datetime.now().date()
                task.completed_date = datetime.now().date()
            elif form.cleaned_data.get('progress_status') == 'In-Progress':
                task.started_date = datetime.now().date() 
            task.last_updated_by_id=request.user.id
            task.save()
        else:
            print("Invalid Form")
            print(form.errors)

        return redirect('home')
    else:
        form=NewTaskForm(user=request.user.id,role=request.user.role)
        return render(request, 'auth/widgets/new_task.html',{'form':form})

def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            user.save()
            raw_password = form.cleaned_data.get('password1')
            user = authenticate(username=user.username, password=raw_password)
            login(request, user)
            return redirect('home')
    else:
        form = SignUpForm()
    return render(request, 'auth/signup.html', { 'form' : form })